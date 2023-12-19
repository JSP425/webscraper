import webscraper as ws
import os
import pytest
from openpyxl import load_workbook # to test formatting in Excel

exDir = "C:/Users/jpark/Desktop/"
clTestAddress = ["Vancouver"]

# @pytest.mark.skip(reason="temporarily skipping")
def test_report(random_name):
    inst = ws.report(exDir + random_name + ".xlsx")

    inst.createSearchTab("testTab", searchTerm="aa", cl = clTestAddress)
    inst.workbook.close() # <-- needed to create the workbook but will cause warning; otherwise load_workbook cannot find it
    testWB = load_workbook(exDir + random_name + ".xlsx")
    testWS = testWB["testTab"]

    inst.run()

    targetTestCells = ["B1", "D1", "F1", "I1"]

    assert isinstance(inst, ws.report)

    assert os.path.exists(exDir + random_name + ".xlsx") == True

    for eachCell in targetTestCells:
        assert testWS[eachCell].font.bold == True
        assert testWS[eachCell].number_format == 'General'
    assert 59.0 <= testWS.column_dimensions['D'].width <= 61.0  # Check width; due to rounding, will not return 60 exactly

    os.remove(exDir + random_name + ".xlsx")

def test_createSearchTab_empty_input(random_name):
    inst = ws.report(exDir + random_name + ".xlsx")

    # Test case: Invalid tab name (empty string)
    with pytest.raises(ValueError, match="Tab name cannot be empty"):
        inst.createSearchTab("")

@pytest.mark.skip(reason="does not apply to dl searches. Need to reconfigure")
def test_createSearchTab_invalid_input(random_name):
    inst = ws.report(exDir + random_name + ".xlsx")

    # Test case: Invalid search term
    with pytest.raises(ValueError, match="Invalid search term"):
        inst.createSearchTab("testTab", searchTerm="", cl = clTestAddress)

def test_createSearchTab_no_place(random_name):
    inst = ws.report(exDir + random_name + ".xlsx")

    with pytest.raises(ValueError, match ="No places to search"):
        inst.createSearchTab("testTab", searchTerm = "test")

def test_createSearchTab_one_place(random_name):
    inst = ws.report(exDir + random_name + ".xlsx")

    try:
        inst.createSearchTab("testTab", searchTerm="test", cl=clTestAddress)
    except ValueError:
        pytest.fail("ValueError should not be raised when there is data in cl")

